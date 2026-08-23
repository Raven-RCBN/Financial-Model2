from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from seed_report_tables import COMPANY_ID, DB, PROJECT_ID, WORKBOOK, extract_grid


ANALYSIS = Path(__file__).resolve().parents[1] / "public" / "workbook-analysis.json"


def main():
    database = json.loads(DB.read_text())
    analysis = json.loads(ANALYSIS.read_text())
    values_wb = load_workbook(WORKBOOK, data_only=True, read_only=False)
    styles_wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    input_sheet_names = (
        analysis["userInputSheets"]
        + analysis["masterDataEntrySheets"]
        + analysis["embeddedSheetInputSheets"]
    )
    input_tables = []
    for sheet_name in input_sheet_names:
        table = extract_grid(values_wb, styles_wb, sheet_name)
        table["id"] = f"input_table_{table['id'].removeprefix('report_table_')}"
        table["companyId"] = COMPANY_ID
        table["projectId"] = PROJECT_ID
        table["sheetName"] = sheet_name
        table["sourceRole"] = "input_screen"
        input_tables.append(table)

    database["inputTables"] = input_tables
    if not any(index["name"] == "input_tables_project_sheet" for index in database["indexes"]):
        database["indexes"].append(
            {
                "name": "input_tables_project_sheet",
                "collection": "inputTables",
                "fields": ["companyId", "projectId", "sheetName"],
            }
        )
    DB.write_text(json.dumps(database, indent=2) + "\n")
    print(
        json.dumps(
            {
                "inputTables": [
                    {
                        "sheetName": table["sheetName"],
                        "rows": table["rowCount"],
                        "columns": table["columnCount"],
                        "range": table["range"],
                    }
                    for table in input_tables
                ]
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
