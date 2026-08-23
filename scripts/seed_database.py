from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/admin/Desktop/OPSL_Budget Requisition_Aug26_FINAL_10Aug26.xlsm")
ANALYSIS = ROOT / "public" / "workbook-analysis.json"
DB = ROOT / "data" / "plantation-financial-model.db.json"

COMPANY_ID = "company_opsl"
PROJECT_ID = "project_opsl_15000ha_development"
MODEL_ID = "model_opsl_aug26"


def cell(ws, address):
    value = ws[address].value
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def scalar(value):
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def normalize_number(value):
    if isinstance(value, (int, float)):
        return value
    return None


def collect_status_transactions(wb):
    ws = wb["STATUS"]
    records = []
    for row in range(12, min(ws.max_row, 90) + 1):
        bank = ws.cell(row, 1).value
        category = ws.cell(row, 2).value
        item = ws.cell(row, 3).value
        code = ws.cell(row, 4).value
        description = ws.cell(row, 5).value
        transferred = normalize_number(ws.cell(row, 6).value)
        not_transferred = normalize_number(ws.cell(row, 7).value)
        notes = ws.cell(row, 8).value
        amount = transferred if transferred is not None else not_transferred
        if not description or amount is None:
            continue
        records.append(
            {
                "id": f"txn_status_{row}",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sourceSheet": "STATUS",
                "sourceRow": row,
                "bankAccount": scalar(bank),
                "category": scalar(category),
                "item": scalar(item),
                "accountCode": str(code) if code is not None else "",
                "description": scalar(description),
                "amountUsd": amount,
                "status": "transferred" if transferred is not None else "not_transferred",
                "notes": scalar(notes),
            }
        )
    return records


def collect_budget_request_transactions(wb):
    ws = wb["OPSL AUG BUD req"]
    records = []
    section = ""
    for row in range(10, min(ws.max_row, 150) + 1):
        col_b = ws.cell(row, 2).value
        col_c = normalize_number(ws.cell(row, 3).value)
        col_e = normalize_number(ws.cell(row, 5).value)
        col_h = normalize_number(ws.cell(row, 8).value)
        if isinstance(col_b, str) and col_b.strip() and col_c is None and col_e is None:
            section = col_b.strip()
        if not isinstance(col_b, str) or not col_b.strip():
            continue
        if col_c is None and col_e is None and col_h is None:
            continue
        records.append(
            {
                "id": f"txn_augreq_{row}",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sourceSheet": "OPSL AUG BUD req",
                "sourceRow": row,
                "category": section,
                "description": col_b.strip(),
                "requestedAmountUsd": col_c or 0,
                "augBudgetUsd": col_e or 0,
                "totalAvailableBudgetUsd": col_h or 0,
                "status": "draft_request",
            }
        )
    return records


def slug(value):
    return (
        str(value)
        .lower()
        .replace(" ", "_")
        .replace("&", "and")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "_")
        .replace("/", "_")
        .replace("$", "usd")
    )


def collect_price_input_records(wb):
    records = []
    cpo = wb["CPO Prices"]
    for year_col in range(1, 20, 3):
        year = cpo.cell(3, year_col).value
        currency = cpo.cell(3, year_col + 1).value
        if not isinstance(year, int):
            continue
        for row in range(4, 17):
            month = cpo.cell(row, year_col).value
            price = cpo.cell(row, year_col + 1).value
            if month in (None, "") or price in (None, ""):
                continue
            records.append(
                {
                    "id": f"input_cpo_price_{year}_{slug(month)}",
                    "companyId": COMPANY_ID,
                    "projectId": PROJECT_ID,
                    "sheetName": "CPO Prices",
                    "cell": cpo.cell(row, year_col + 1).coordinate,
                    "label": f"CPO {year} {month} price ({currency})",
                    "value": price,
                    "entryType": "master_data",
                    "status": "active",
                    "source": "seeded_from_workbook",
                }
            )
    records.extend(
        [
            {
                "id": "input_cpo_2025_average_price_rm",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sheetName": "CPO Prices",
                "cell": "A24",
                "label": "Average CPO Rotterdam selling price 2025 (RM)",
                "value": cpo["A24"].value,
                "entryType": "master_data",
                "status": "active",
                "source": "seeded_from_workbook",
            },
            {
                "id": "input_cpo_model_price_usd",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sheetName": "CPO Prices",
                "cell": "B24",
                "label": "CPO price used in model (USD)",
                "value": cpo["B24"].value,
                "entryType": "master_data",
                "status": "active",
                "source": "seeded_from_workbook",
                "reportLinks": [
                    {
                        "sheetName": "Financials",
                        "type": "scheduleEscalatedRow",
                        "label": "CPO Market Price",
                        "movementLabel": "CPO price movement",
                    }
                ],
            },
        ]
    )

    pk = wb["PK Prices"]
    for year_col in (2, 5, 8):
        year = pk.cell(3, year_col).value
        currency = pk.cell(3, year_col + 1).value
        if not isinstance(year, int):
            continue
        for row in range(4, 17):
            month = pk.cell(row, year_col).value
            price = pk.cell(row, year_col + 1).value
            if month in (None, "") or price in (None, ""):
                continue
            records.append(
                {
                    "id": f"input_pk_price_{year}_{slug(month)}",
                    "companyId": COMPANY_ID,
                    "projectId": PROJECT_ID,
                    "sheetName": "PK Prices",
                    "cell": pk.cell(row, year_col + 1).coordinate,
                    "label": f"PK {year} {month} price ({currency})",
                    "value": price,
                    "entryType": "master_data",
                    "status": "active",
                    "source": "seeded_from_workbook",
                }
            )
    records.extend(
        [
            {
                "id": "input_pk_average_price_naira",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sheetName": "PK Prices",
                "cell": "K4",
                "label": "PK average price (Naira)",
                "value": pk["K4"].value,
                "entryType": "master_data",
                "status": "active",
                "source": "seeded_from_workbook",
            },
            {
                "id": "input_pk_model_price_2024_naira",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sheetName": "PK Prices",
                "cell": "L4",
                "label": "PK price used in model 2024 (Naira)",
                "value": pk["L4"].value,
                "entryType": "master_data",
                "status": "active",
                "source": "seeded_from_workbook",
            },
        ]
    )
    return records


def build_input_records(analysis, metrics=None, wb=None):
    records = [
        {
            "id": "input_reporting_currency",
            "companyId": COMPANY_ID,
            "projectId": PROJECT_ID,
            "sheetName": "Reporting Settings",
            "cell": "Master",
            "label": "Reporting currency",
            "value": "USD",
            "settingKey": "reportingCurrency",
            "entryType": "reporting_setting",
            "status": "active",
            "source": "master_data_setting",
        }
    ]
    metrics = metrics or {}
    metric_labels = {
        "totalDevelopmentExpenditure": "Total development expenditure",
        "committedSources": "Committed funding sources",
        "costPerHa": "Cost per hectare",
        "nominalAfterTaxIrr": "Nominal after-tax IRR",
        "paybackYears": "Payback period",
        "nominalAfterTaxNpvAtWacc": "NPV @ WACC",
        "year1BudgetAmount": "Year 1 budget amount",
        "netBudgetAvailable": "Net budget available",
        "augCurrentCashBalance": "August current cash balance",
    }
    for metric_key, label in metric_labels.items():
        records.append(
            {
                "id": f"input_metric_{metric_key}",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sheetName": "Gen Inputs",
                "cell": "Metric",
                "label": label,
                "value": metrics.get(metric_key, ""),
                "metricKey": metric_key,
                "entryType": "model_metric",
                "status": "active",
                "source": "seeded_from_workbook",
            }
        )
    entry_sheets = (
        analysis["userInputSheets"]
        + analysis["masterDataEntrySheets"]
        + analysis["embeddedSheetInputSheets"]
    )
    by_name = {sheet["name"]: sheet for sheet in analysis["sheets"]}
    for sheet_name in entry_sheets:
        if sheet_name in {"CPO Prices", "PK Prices"}:
            continue
        sheet = by_name[sheet_name]
        samples = sheet.get("yellowInputs") or [
            {"cell": "TABLE", "label": f"{sheet_name} mapped table", "value": "Database records"}
        ]
        for index, item in enumerate(samples):
            records.append(
                {
                    "id": f"input_{sheet_name.lower().replace(' ', '_').replace('&', 'and').replace('(', '').replace(')', '')}_{index + 1}",
                    "companyId": COMPANY_ID,
                    "projectId": PROJECT_ID,
                    "sheetName": sheet_name,
                    "cell": item.get("cell", ""),
                    "label": item.get("label") or "Input field",
                    "value": item.get("value", ""),
                    "entryType": sheet.get("role"),
                    "status": "active",
                    "source": "seeded_from_workbook",
                }
            )
    if wb:
        records.extend(collect_price_input_records(wb))
    return records


def build_formula_rules(analysis):
    rules = []
    for sheet in analysis["sheets"]:
        if not sheet.get("inSystemScope"):
            continue
        for item in sheet.get("sampleFormulas", []):
            rules.append(
                {
                    "id": f"formula_{sheet['name'].lower().replace(' ', '_').replace('&', 'and').replace('(', '').replace(')', '')}_{item['cell'].lower()}",
                    "companyId": COMPANY_ID,
                    "projectId": PROJECT_ID,
                    "sheetName": sheet["name"],
                    "cell": item["cell"],
                    "formula": item["formula"],
                    "status": "approved_baseline",
                    "editable": True,
                    "source": "seeded_from_workbook",
                }
            )
    return rules


def build_report_snapshots(analysis):
    excluded = set(analysis["userInputSheets"] + analysis["masterDataEntrySheets"])
    snapshots = []
    by_name = {sheet["name"]: sheet for sheet in analysis["sheets"]}
    for name in analysis["systemSheetOrderLeftToRight"]:
        if name in excluded:
            continue
        sheet = by_name[name]
        snapshots.append(
            {
                "id": f"report_{name.lower().replace(' ', '_').replace('&', 'and').replace('(', '').replace(')', '')}",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sheetName": name,
                "dimensions": sheet["dimensions"],
                "formulaCount": sheet["formulaCount"],
                "contextualInputCount": sheet["yellowInputCount"],
                "headings": sheet.get("headings", []),
                "sampleFormulas": sheet.get("sampleFormulas", []),
                "status": "baseline_migrated",
            }
        )
    return snapshots


def main():
    analysis = json.loads(ANALYSIS.read_text())
    values_wb = load_workbook(WORKBOOK, data_only=True, read_only=True)

    summary = values_wb["Summary (US$)"]
    status = values_wb["STATUS"]
    aug_request = values_wb["OPSL AUG BUD req"]

    metrics = {
        "totalDevelopmentExpenditure": cell(summary, "F17"),
        "committedSources": cell(summary, "I10"),
        "fundingShortfall": cell(summary, "I14"),
        "costPerHa": cell(summary, "F29"),
        "nominalAfterTaxIrr": cell(summary, "I24"),
        "paybackYears": cell(summary, "I26"),
        "nominalAfterTaxNpvAtWacc": cell(summary, "I27"),
        "year1BudgetAmount": cell(status, "F5"),
        "netBudgetAvailable": cell(status, "F9"),
        "augCurrentCashBalance": cell(aug_request, "E8"),
    }

    now = datetime.now(timezone.utc).isoformat()
    database = {
        "schemaVersion": 1,
        "seededAt": now,
        "sourceWorkbook": str(WORKBOOK),
        "indexes": [
            {"name": "companies_id", "collection": "companies", "fields": ["id"], "unique": True},
            {"name": "projects_company_id", "collection": "projects", "fields": ["companyId", "id"], "unique": True},
            {"name": "models_project_id", "collection": "models", "fields": ["companyId", "projectId", "id"], "unique": True},
            {"name": "inputs_project_sheet_status", "collection": "inputRecords", "fields": ["companyId", "projectId", "sheetName", "status"]},
            {"name": "formulas_project_sheet_status", "collection": "formulaRules", "fields": ["companyId", "projectId", "sheetName", "status"]},
            {"name": "reports_project_sheet", "collection": "reportSnapshots", "fields": ["companyId", "projectId", "sheetName"]},
            {"name": "transactions_project_category_status", "collection": "transactions", "fields": ["companyId", "projectId", "category", "status"]},
            {"name": "market_project_source", "collection": "marketData", "fields": ["companyId", "projectId", "sourceSheet"]},
        ],
        "companies": [
            {
                "id": COMPANY_ID,
                "name": "Octavus Plantation Ltd",
                "tenant": "Agrilntel portfolio",
                "currency": "USD",
                "status": "active",
            }
        ],
        "projects": [
            {
                "id": PROJECT_ID,
                "companyId": COMPANY_ID,
                "modelId": MODEL_ID,
                "name": "OP Sierra Leone",
                "description": "15,000 Ha new plantation financial model",
                "hectares": 15000,
                "stage": "development",
                "status": "baseline_seeded",
                "settings": {
                    "sourceCurrency": "USD",
                    "reportingCurrency": "USD",
                    "startYear": 2026,
                    "crop": "Oil Palm",
                    "projectionYears": 27,
                    "reportBasis": "Financial Projections",
                    "reportingStandard": "IFRS",
                    "reportHeaderTemplate": "{company} - {project} - {hectares} Ha",
                    "supportedReportingCurrencies": ["USD", "MYR"],
                    "currencyRates": {
                        "USD": 1,
                        "MYR": 4.0845,
                    },
                },
                "sourceWorkbook": WORKBOOK.name,
            }
        ],
        "models": [
            {
                "id": MODEL_ID,
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "name": "Plantation Financial Model - Aug 2026",
                "basis": "IFRS-style management reporting baseline seeded from workbook",
                "period": "Aug 2026",
                "metrics": metrics,
            }
        ],
        "inputRecords": build_input_records(analysis, metrics, values_wb),
        "formulaRules": build_formula_rules(analysis),
        "reportSnapshots": build_report_snapshots(analysis),
        "transactions": collect_status_transactions(values_wb)
        + collect_budget_request_transactions(values_wb),
        "marketData": [
            {
                "id": "market_cpo_prices",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sourceSheet": "CPO Prices",
                "label": "CPO prices ticker",
                "status": "baseline_seeded",
            },
            {
                "id": "market_pk_prices",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sourceSheet": "PK Prices",
                "label": "PK and PKO commodity prices",
                "status": "baseline_seeded",
            },
            {
                "id": "market_wacc",
                "companyId": COMPANY_ID,
                "projectId": PROJECT_ID,
                "sourceSheet": "WACC Benchmarking",
                "label": "WACC benchmarking and discount-rate assumptions",
                "status": "baseline_seeded",
            },
        ],
    }

    DB.parent.mkdir(parents=True, exist_ok=True)
    DB.write_text(json.dumps(database, indent=2, default=str) + "\n")
    print(
        json.dumps(
            {
                "database": str(DB),
                "companies": len(database["companies"]),
                "projects": len(database["projects"]),
                "inputRecords": len(database["inputRecords"]),
                "formulaRules": len(database["formulaRules"]),
                "reportSnapshots": len(database["reportSnapshots"]),
                "transactions": len(database["transactions"]),
                "metrics": metrics,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
