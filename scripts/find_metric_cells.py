from openpyxl import load_workbook

WORKBOOK = "/Users/admin/Desktop/OPSL_Budget Requisition_Aug26_FINAL_10Aug26.xlsm"

targets = [
    "Total",
    "Total Committed Sources",
    "Excess / (Shortfall)",
    "Nominal After-Tax IRR",
    "Payback period",
    "Valuation using NPV",
    "1st year budget amount ",
    "TOTAL NET BUDGET AVAILABLE",
    "TOTAL",
]

wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
for sheet_name in ["Summary (US$)", "STATUS", "OPSL AUG BUD req"]:
    ws = wb[sheet_name]
    print(f"--- {sheet_name}")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in targets:
                values = []
                for col in range(cell.column, min(ws.max_column, cell.column + 8) + 1):
                    value = ws.cell(cell.row, col).value
                    if value not in (None, ""):
                        values.append(f"{ws.cell(cell.row, col).coordinate}={value}")
                print(" | ".join(values))
