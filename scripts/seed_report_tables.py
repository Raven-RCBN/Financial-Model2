from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/admin/Desktop/OPSL_Budget Requisition_Aug26_FINAL_10Aug26.xlsm")
DB = ROOT / "data" / "plantation-financial-model.db.json"

COMPANY_ID = "company_opsl"
PROJECT_ID = "project_opsl_15000ha_development"

THEME_COLOR_ORDER = [
    "lt1",
    "dk1",
    "lt2",
    "dk2",
    "accent1",
    "accent2",
    "accent3",
    "accent4",
    "accent5",
    "accent6",
    "hlink",
    "folHlink",
]

SCHEDULE_CONFIGS = {
    "Financials": {
        "titleCell": (1, 1),
        "subtitleCell": (2, 1),
        "reportTitleCell": (3, 1),
        "periodRow": 1,
        "startRow": 2,
        "endRow": 3,
        "sectionCol": 3,
        "labelCol": 4,
        "percentCol": 6,
        "totalCol": 7,
        "firstPeriodCol": 8,
        "lastPeriodCol": 34,
        "firstDataRow": 6,
        "lastDataRow": 97,
    },
    "Detail Cashflow": {
        "titleCell": (1, 1),
        "subtitleCell": (2, 1),
        "reportTitleCell": (2, 1),
        "periodRow": 5,
        "startRow": 6,
        "endRow": 7,
        "sectionCol": 2,
        "labelCol": 2,
        "totalCol": 3,
        "firstPeriodCol": 4,
        "lastPeriodCol": 30,
        "firstDataRow": 8,
        "lastDataRow": 70,
    },
    "Balance Sheet": {
        "titleCell": (1, 2),
        "subtitleCell": (2, 2),
        "reportTitleCell": (3, 2),
        "periodRow": 0,
        "startRow": 4,
        "endRow": 4,
        "sectionCol": 2,
        "labelCol": 2,
        "firstPeriodCol": 6,
        "lastPeriodCol": 32,
        "firstDataRow": 5,
        "lastDataRow": 88,
    },
    "Valuation": {
        "titleCell": (1, 1),
        "subtitleCell": (2, 1),
        "reportTitleCell": (3, 1),
        "periodRow": 1,
        "startRow": 2,
        "endRow": 3,
        "sectionCol": 4,
        "labelCol": 4,
        "labelCols": [4, 5],
        "totalCol": 8,
        "firstPeriodCol": 9,
        "lastPeriodCol": 35,
        "firstDataRow": 6,
        "lastDataRow": 132,
        "rowLabels": {
            81: "Total project cash flow",
            108: "ROI (revenue)",
        },
        "skipRows": [92],
    },
}

BUDGET_GRID_CONFIGS = {
    "Year 1 Budget_OPSL": {
        "titleCell": (1, 1),
        "subtitleCell": (2, 1),
        "reportTitleCell": (2, 1),
        "firstDataRow": 4,
        "firstLabelCol": 2,
        "columnOrder": [
            ("line_item", "Line item", 2),
            ("code", "Code", 1),
            ("total", "Total", 3),
            ("year_1", "Year 1", 4),
            ("unit_cost", "Unit cost", 5),
            ("units", "No. of units", 6),
            ("month_1", "Jun 2026", 7),
            ("month_2", "Jul 2026", 8),
            ("month_3", "Aug 2026", 9),
            ("month_4", "Sep 2026", 10),
            ("month_5", "Oct 2026", 11),
            ("month_6", "Nov 2026", 12),
            ("month_7", "Dec 2026", 13),
            ("month_8", "Jan 2027", 14),
            ("month_9", "Feb 2027", 15),
            ("month_10", "Mar 2027", 16),
            ("month_11", "Apr 2027", 17),
            ("month_12", "May 2027", 18),
            ("check", "Check", 19),
        ],
    },
    "OPSL AUG BUD req": {
        "titleCell": (1, 1),
        "subtitleCell": (3, 1),
        "reportTitleCell": (6, 1),
        "headerRow": 10,
        "firstDataRow": 11,
        "firstLabelCol": 2,
    },
    "Fund Req Aug26": {
        "titleCell": (1, 2),
        "subtitleCell": (3, 2),
        "reportTitleCell": (1, 2),
        "headerRow": 5,
        "firstDataRow": 6,
        "firstLabelCol": 2,
    },
    "3-Mth Budget": {
        "titleCell": (1, 2),
        "subtitleCell": (3, 2),
        "reportTitleCell": (1, 2),
        "headerRow": 5,
        "firstDataRow": 6,
        "firstLabelCol": 2,
    },
}

BUDGET_GRID_CONFIGS["Year 1 Budget"] = {
    **BUDGET_GRID_CONFIGS["Year 1 Budget_OPSL"],
    "reportTitleCell": (2, 1),
}

WORKSHEET_GRID_CONFIGS = {
    "Summary (US$)": {
        "firstDataRow": 6,
        "titleCell": (1, 1),
        "subtitleCell": (2, 1),
        "reportTitleCell": (3, 1),
    }
}


def scalar(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def non_empty(value):
    return value not in (None, "", " ")


def is_source_text(value):
    return isinstance(value, str) and value.strip().lower().startswith("source")


def label_from_row(ws, row, config):
    if row in config.get("rowLabels", {}):
        return config["rowLabels"][row], config["labelCol"]
    label_cols = config.get("labelCols") or list(
        range(config["labelCol"], max(config["labelCol"], config["firstPeriodCol"] - 1) + 1)
    )
    for col in label_cols:
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip():
            return value.strip(), col
    return "", config["labelCol"]


def is_blank_schedule_row(label, total, values):
    if label:
        return False
    meaningful_values = [value for value in values if value not in (None, "")]
    if total not in (None, "", 0):
        return False
    if not meaningful_values:
        return True
    if all(value == 0 for value in meaningful_values):
        return True
    if all(clean(value).lower() == "projection" for value in meaningful_values):
        return True
    return False


def meaningful_grid_row(row):
    cells = row.get("cells", [])
    populated = [
        (index, cell.get("value"))
        for index, cell in enumerate(cells)
        if non_empty(cell.get("value"))
    ]
    if not populated:
        return False
    if any(is_source_text(value) for _, value in populated):
        return True
    first_index = populated[0][0]
    first_value = populated[0][1]
    if first_index <= 7 and not (is_number(first_value) and first_value == 0 and len(populated) == 1):
        return True
    if len(populated) >= 3:
        return True
    if any(isinstance(value, str) and value.strip() for _, value in populated):
        return first_index <= 10
    return False


def trim_grid_tail(rows):
    end = len(rows)
    while end > 0 and not meaningful_grid_row(rows[end - 1]):
        end -= 1
    return rows[:end]


def compact_empty_columns(columns, rows):
    keep_indexes = []
    for index, _column in enumerate(columns):
        if any(
            index < len(row.get("cells", [])) and non_empty(row["cells"][index].get("value"))
            for row in rows
        ):
            keep_indexes.append(index)
    if len(keep_indexes) == len(columns):
        return columns, rows
    compacted_columns = [columns[index] for index in keep_indexes]
    compacted_rows = []
    for row in rows:
        next_row = dict(row)
        cells = row.get("cells", [])
        next_row["cells"] = [cells[index] for index in keep_indexes if index < len(cells)]
        compacted_rows.append(next_row)
    return compacted_columns, compacted_rows


def sheet_slug(name):
    return (
        name.lower()
        .replace(" ", "_")
        .replace("&", "and")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "_")
    )


def workbook_theme_colors(workbook):
    if not workbook.loaded_theme:
        return {}
    root = ET.fromstring(workbook.loaded_theme)
    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    scheme = root.find(".//a:clrScheme", namespace)
    colors = {}
    if scheme is None:
        return colors
    for index, child in enumerate(list(scheme)):
        key = child.tag.split("}", 1)[-1]
        srgb = child.find(".//a:srgbClr", namespace)
        system = child.find(".//a:sysClr", namespace)
        value = None
        if srgb is not None:
            value = srgb.attrib.get("val")
        elif system is not None:
            value = system.attrib.get("lastClr")
        if value:
            colors[key] = value.upper()
    return colors


THEME_COLORS = {}


def apply_tint(rgb, tint):
    if not rgb or tint in (None, 0, 0.0):
        return rgb
    channels = [int(rgb[index : index + 2], 16) for index in (0, 2, 4)]
    adjusted = []
    for channel in channels:
        if tint < 0:
            value = channel * (1 + tint)
        else:
            value = channel * (1 - tint) + (255 * tint)
        adjusted.append(max(0, min(255, round(value))))
    return "".join(f"{channel:02X}" for channel in adjusted)


def resolve_color(color):
    if not color:
        return ""
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    if color.type == "theme" and color.theme is not None:
        theme_key = THEME_COLOR_ORDER[color.theme] if color.theme < len(THEME_COLOR_ORDER) else str(color.theme)
        rgb = THEME_COLORS.get(theme_key, "")
        return apply_tint(rgb, color.tint or 0)
    if color.type == "indexed" and color.indexed is not None:
        indexed = {
            0: "000000",
            1: "FFFFFF",
            2: "FF0000",
            3: "00FF00",
            4: "0000FF",
            5: "FFFF00",
            6: "FF00FF",
            7: "00FFFF",
            64: "",
        }
        return indexed.get(color.indexed, "")
    return ""


def fill_rgb(cell):
    if not cell.fill or not cell.fill.fill_type:
        return ""
    return resolve_color(cell.fill.fgColor)


def font_rgb(cell):
    if not cell.font or not cell.font.color:
        return ""
    return resolve_color(cell.font.color)


def excel_style(cell):
    style = {}
    fill = fill_rgb(cell)
    font = font_rgb(cell)
    if fill and fill != "FFFFFF":
        style["fillColor"] = f"#{fill}"
    if font and (font != "000000" or fill not in ("", "FFFFFF")):
        style["fontColor"] = f"#{font}"
    return style


def is_dark_fill(rgb):
    if not rgb:
        return False
    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) < 100


def cell_style(cell, value):
    rgb = fill_rgb(cell)
    style = []
    if cell.font.bold:
        style.append("bold")
    if rgb and rgb != "FFFFFF":
        if rgb in {"002060", "1F4E79", "17365D", "244062"} or is_dark_fill(rgb):
            style.append("dark")
        elif rgb.startswith("FF") or rgb in {"FFC000", "FFFF00", "FFD966"}:
            style.append("yellow")
        elif rgb in {"DCECC6", "E2F0D9", "C6E0B4", "A9D18E"}:
            style.append("green")
        else:
            style.append("fill")
    if is_number(value):
        style.append("number")
        if value < 0:
            style.append("negative")
    if cell.alignment.horizontal == "center":
        style.append("center")
    return " ".join(style)


def row_style_from_cells(cells):
    styles = [cell.get("style", "") for cell in cells]
    dark_count = sum("dark" in style for style in styles)
    filled_count = sum(any(token in style for token in ("dark", "green", "yellow", "fill")) for style in styles)
    if dark_count >= max(2, len(styles) // 2):
        return "dark"
    if any("green" in style for style in styles) and filled_count >= max(2, len(styles) // 3):
        return "section"
    if any("yellow" in style for style in styles) and filled_count >= max(2, len(styles) // 3):
        return "yellow"
    if any("bold" in style for style in styles):
        return "subheader"
    return "line"


def column_label(value, fallback):
    label = clean(scalar(value))
    return label if label else fallback


def grid_cell(values_ws, styles_ws, row_number, col):
    value = values_ws.cell(row_number, col).value
    style_cell = styles_ws.cell(row_number, col)
    cell = {
        "address": f"{get_column_letter(col)}{row_number}",
        "value": scalar(value),
        "style": cell_style(style_cell, value),
    }
    style = excel_style(style_cell)
    if style:
        cell["excelStyle"] = style
    return cell


def grid_has_meaningful_label(cells, first_label_index=0):
    if not cells:
        return False
    label_value = cells[first_label_index].get("value") if first_label_index < len(cells) else ""
    if non_empty(label_value):
        return True
    return any(non_empty(cell.get("value")) for cell in cells)


def extract_configured_grid(values_wb, styles_wb, sheet_name, config):
    ws = values_wb[sheet_name]
    style_ws = styles_wb[sheet_name]
    min_row, min_col, max_row, max_col = effective_bounds(ws)
    header_row = config.get("headerRow")
    first_data_row = config.get("firstDataRow", min_row)
    first_label_col = config.get("firstLabelCol", min_col)

    if "columnOrder" in config:
        source_columns = config["columnOrder"]
    else:
        source_columns = [
            (
                f"col_{col}",
                column_label(ws.cell(header_row, col).value if header_row else None, get_column_letter(col)),
                col,
            )
            for col in range(min_col, max_col + 1)
        ]

    rows = []
    first_label_index = next(
        (index for index, (_, _, col) in enumerate(source_columns) if col == first_label_col),
        0,
    )
    for row_number in range(first_data_row, max_row + 1):
        cells = [grid_cell(ws, style_ws, row_number, col) for _, _, col in source_columns]
        if not grid_has_meaningful_label(cells, first_label_index):
            continue
        rows.append(
            {
                "sourceRow": row_number,
                "style": row_style_from_cells(cells),
                "cells": cells,
            }
        )

    rows = trim_grid_tail(rows)
    columns = [
        {"key": key, "label": label, "sourceColumn": col}
        for key, label, col in source_columns
    ]
    columns, rows = compact_empty_columns(columns, rows)
    title_row, title_col = config.get("titleCell", (min_row, min_col))
    subtitle_row, subtitle_col = config.get("subtitleCell", (title_row + 1, title_col))
    report_row, report_col = config.get("reportTitleCell", (title_row, title_col))
    return {
        "id": f"report_table_{sheet_slug(sheet_name)}",
        "kind": "worksheet-grid",
        "presentation": "budget-grid",
        "companyId": COMPANY_ID,
        "projectId": PROJECT_ID,
        "sheetName": sheet_name,
        "title": clean(ws.cell(title_row, title_col).value) or sheet_name,
        "subtitle": clean(ws.cell(subtitle_row, subtitle_col).value) or "",
        "reportTitle": clean(ws.cell(report_row, report_col).value) or sheet_name,
        "titleStyle": excel_style(style_ws.cell(report_row, report_col)),
        "currency": "USD",
        "range": f"{get_column_letter(min_col)}{first_data_row}:{get_column_letter(max_col)}{max_row}",
        "columnCount": len(columns),
        "rowCount": len(rows),
        "columns": columns,
        "rows": rows,
    }


def style_for_schedule(ws, row, label_col, values):
    label = clean(ws.cell(row, label_col).value)
    section_label = clean(ws.cell(row, max(1, label_col - 1)).value)
    rgb = fill_rgb(ws.cell(row, label_col))
    has_values = any(is_number(value) for value in values)
    bold = bool(ws.cell(row, label_col).font.bold)
    upper = label.upper()
    if section_label and not label:
        return "section"
    if rgb == "002060":
        return "blue-total"
    if upper.startswith("TOTAL") or upper in {
        "EBITDA",
        "OPERATING PROFIT",
        "PROFIT BEFORE TAX",
        "PROFIT AFTER TAX",
        "CASH FLOW FROM OPERATIONS",
        "CASH FLOW AVAILABLE FOR DIVIDENDS",
        "ENDING CASH BALANCE C/F",
        "TOTAL RECEIPTS",
        "TOTAL PAYMENTS",
        "TOTAL NCA",
        "TOTAL CA",
        "TOTAL LIABILITIES",
    }:
        return "total"
    if bold and not has_values:
        return "subheader"
    return "line"


def schedule_cell_styles(style_ws, row, config, label_col, periods):
    styles = {
        "label": excel_style(style_ws.cell(row, label_col)),
        "values": [excel_style(style_ws.cell(row, period["column"])) for period in periods],
    }
    if config.get("percentCol"):
        styles["percent"] = excel_style(style_ws.cell(row, config["percentCol"]))
    if config.get("totalCol"):
        styles["total"] = excel_style(style_ws.cell(row, config["totalCol"]))
    compact = {key: value for key, value in styles.items() if key != "values" and value}
    if any(styles["values"]):
        compact["values"] = styles["values"]
    return compact


def schedule_header_styles(style_ws, config, periods):
    period_row = config.get("periodRow") or config.get("startRow") or config.get("firstDataRow")
    year_row = config.get("endRow") or period_row
    label_col = config["labelCol"]
    styles = {
        "period": {
            "label": excel_style(style_ws.cell(period_row, label_col)),
            "percent": excel_style(style_ws.cell(period_row, config["percentCol"])) if config.get("percentCol") else {},
            "total": excel_style(style_ws.cell(period_row, config["totalCol"])) if config.get("totalCol") else {},
            "values": [excel_style(style_ws.cell(period_row, period["column"])) for period in periods],
        },
        "year": {
            "label": excel_style(style_ws.cell(year_row, label_col)),
            "percent": excel_style(style_ws.cell(year_row, config["percentCol"])) if config.get("percentCol") else {},
            "total": excel_style(style_ws.cell(year_row, config["totalCol"])) if config.get("totalCol") else {},
            "values": [excel_style(style_ws.cell(year_row, period["column"])) for period in periods],
        },
    }
    compact = {}
    for row_key, row_styles in styles.items():
        row_compact = {key: value for key, value in row_styles.items() if key != "values" and value}
        if any(row_styles["values"]):
            row_compact["values"] = row_styles["values"]
        if row_compact:
            compact[row_key] = row_compact
    return compact


def extract_periods(ws, period_row, start_row, end_row, first_col, last_col):
    periods = []
    for col in range(first_col, last_col + 1):
        raw_period = ws.cell(period_row, col).value if period_row else len(periods)
        if raw_period in (None, ""):
            raw_period = len(periods)
        end_value = ws.cell(end_row, col).value if end_row else None
        periods.append(
            {
                "key": f"p{len(periods)}",
                "column": col,
                "period": int(raw_period) if is_number(raw_period) else len(periods),
                "startDate": scalar(ws.cell(start_row, col).value) if start_row else "",
                "endDate": scalar(end_value) if end_row else "",
                "label": str(int(raw_period)) if is_number(raw_period) else f"{len(periods)}",
                "year": str(end_value.year) if isinstance(end_value, datetime) else "",
            }
        )
    return periods


def extract_schedule(values_wb, styles_wb, sheet_name, config):
    ws = values_wb[sheet_name]
    style_ws = styles_wb[sheet_name]
    periods = extract_periods(
        ws,
        config["periodRow"],
        config.get("startRow"),
        config.get("endRow"),
        config["firstPeriodCol"],
        config["lastPeriodCol"],
    )
    rows = []
    current_section = ""
    for row in range(config["firstDataRow"], config["lastDataRow"] + 1):
        if row in config.get("skipRows", []):
            continue
        label, actual_label_col = label_from_row(ws, row, config)
        section = clean(ws.cell(row, config.get("sectionCol", config["labelCol"] - 1)).value)
        if section and not label:
            current_section = section
            section_style_col = config.get("sectionCol") or config["labelCol"]
            rows.append(
                {
                    "sourceRow": row,
                    "label": section,
                    "section": section,
                    "kind": "section",
                    "percent": None,
                    "total": None,
                    "values": [],
                    "style": "section",
                    "cellStyles": schedule_cell_styles(style_ws, row, config, section_style_col, []),
                }
            )
            continue
        values = [ws.cell(row, period["column"]).value for period in periods]
        total = ws.cell(row, config["totalCol"]).value if config.get("totalCol") else None
        percent = ws.cell(row, config["percentCol"]).value if config.get("percentCol") else None
        if is_blank_schedule_row(label, total, values):
            continue
        rows.append(
            {
                "sourceRow": row,
                "label": label,
                "section": current_section,
                "kind": "line",
                "percent": scalar(percent),
                "total": scalar(total),
                "values": [scalar(value) for value in values],
                "style": style_for_schedule(style_ws, row, actual_label_col, values),
                "cellStyles": schedule_cell_styles(style_ws, row, config, actual_label_col, periods),
            }
        )

    return {
        "id": f"report_table_{sheet_slug(sheet_name)}",
        "kind": "projection-schedule",
        "companyId": COMPANY_ID,
        "projectId": PROJECT_ID,
        "sheetName": sheet_name,
        "title": scalar(ws.cell(config["titleCell"][0], config["titleCell"][1]).value),
        "subtitle": scalar(ws.cell(config["subtitleCell"][0], config["subtitleCell"][1]).value),
        "reportTitle": scalar(ws.cell(config["reportTitleCell"][0], config["reportTitleCell"][1]).value),
        "titleStyle": excel_style(style_ws.cell(config["reportTitleCell"][0], config["reportTitleCell"][1])),
        "headerStyles": schedule_header_styles(style_ws, config, periods),
        "currency": "USD",
        "periods": periods,
        "columns": {
            "label": "Line item",
            "percent": "%",
            "total": "Total",
        },
        "rows": rows,
    }


def effective_bounds(ws):
    min_row = min_col = 10**9
    max_row = max_col = 0
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value not in (None, ""):
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if max_row == 0:
        return 1, 1, 1, 1
    return min_row, min_col, max_row, max_col


def extract_grid(values_wb, styles_wb, sheet_name, config=None):
    config = config or {}
    ws = values_wb[sheet_name]
    style_ws = styles_wb[sheet_name]
    min_row, min_col, max_row, max_col = effective_bounds(ws)
    first_data_row = max(config.get("firstDataRow", min_row), min_row)
    rows = []
    for row_number in range(first_data_row, max_row + 1):
        cells = []
        has_value = False
        for col in range(min_col, max_col + 1):
            value = ws.cell(row_number, col).value
            has_value = has_value or value not in (None, "")
            cells.append(grid_cell(ws, style_ws, row_number, col))
        if has_value:
            rows.append(
                {
                    "sourceRow": row_number,
                    "style": row_style_from_cells(cells),
                    "cells": cells,
                }
            )
    rows = trim_grid_tail(rows)
    columns = [
        {"key": get_column_letter(col), "label": get_column_letter(col), "sourceColumn": col}
        for col in range(min_col, max_col + 1)
    ]
    columns, rows = compact_empty_columns(columns, rows)
    title_row, title_col = config.get("titleCell", (min_row, min_col))
    subtitle_row, subtitle_col = config.get("subtitleCell", (title_row + 1, title_col))
    report_row, report_col = config.get("reportTitleCell", (title_row, title_col))
    title = clean(ws.cell(title_row, title_col).value) or sheet_name
    subtitle = clean(ws.cell(subtitle_row, subtitle_col).value) if subtitle_row <= max_row else ""
    return {
        "id": f"report_table_{sheet_slug(sheet_name)}",
        "kind": "worksheet-grid",
        "companyId": COMPANY_ID,
        "projectId": PROJECT_ID,
        "sheetName": sheet_name,
        "title": title,
        "subtitle": subtitle,
        "reportTitle": sheet_name,
        "titleStyle": excel_style(style_ws.cell(report_row, report_col)),
        "currency": "USD",
        "range": f"{get_column_letter(min_col)}{first_data_row}:{get_column_letter(max_col)}{max_row}",
        "columnCount": len(columns),
        "rowCount": len(rows),
        "columns": columns,
        "rows": rows,
    }


def main():
    global THEME_COLORS
    values_wb = load_workbook(WORKBOOK, data_only=True, read_only=False)
    styles_wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    THEME_COLORS = workbook_theme_colors(styles_wb)
    database = json.loads(DB.read_text())
    report_names = [report["sheetName"] for report in database.get("reportSnapshots", [])]
    report_tables = []
    for sheet_name in report_names:
        if sheet_name in SCHEDULE_CONFIGS:
            report_tables.append(extract_schedule(values_wb, styles_wb, sheet_name, SCHEDULE_CONFIGS[sheet_name]))
        elif sheet_name in BUDGET_GRID_CONFIGS:
            report_tables.append(extract_configured_grid(values_wb, styles_wb, sheet_name, BUDGET_GRID_CONFIGS[sheet_name]))
        else:
            report_tables.append(extract_grid(values_wb, styles_wb, sheet_name, WORKSHEET_GRID_CONFIGS.get(sheet_name)))

    database["reportTables"] = report_tables
    if not any(index["name"] == "report_tables_project_sheet" for index in database["indexes"]):
        database["indexes"].append(
            {
                "name": "report_tables_project_sheet",
                "collection": "reportTables",
                "fields": ["companyId", "projectId", "sheetName"],
            }
        )
    DB.write_text(json.dumps(database, indent=2) + "\n")
    print(
        json.dumps(
            {
                "reportTables": [
                    {
                        "sheetName": table["sheetName"],
                        "kind": table["kind"],
                        "rows": len(table["rows"]),
                        "columns": len(table.get("periods", table.get("columns", []))),
                    }
                    for table in report_tables
                ]
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
